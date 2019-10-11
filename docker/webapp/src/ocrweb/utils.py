# -*- coding: utf-8 -*-

import os
import zipfile
from datetime import datetime


class Utils(object):

    @classmethod
    def create_xlsx(cls, stroke_name='', orders=None):
        if orders:
            from openpyxl import Workbook
            base_path = os.getcwd()
            zip_filename = '{0}.zip'.format(datetime.now().strftime('%Y%m%d%H%M%S'))
            xls_filename = '{0}.xlsx'.format(stroke_name)
            meter_wb = Workbook()
            wx_title = [stroke_name]
            ws_header = ['序号', '姓名', '物品名称', '收货地址']
            cur_ws = meter_wb.active
            cur_ws.title = '物品核对'
            cur_ws.column_dimensions['C'].width = 40
            cur_ws.column_dimensions['C'].style.alignment.wrap_text = True
            cur_ws.column_dimensions['D'].width = 60
            cur_ws.column_dimensions['D'].style.alignment.wrap_text = True
            cur_ws.merge_cells('A1:D1')
            cur_ws.row_dimensions[1].height = 40
            cur_ws.append(wx_title)
            cur_ws.append(ws_header)
            zip_filepath = os.path.join(base_path, 'tmp', zip_filename)
            xls_filepath = os.path.join(base_path, 'tmp', xls_filename)
            with zipfile.ZipFile(zip_filepath, mode='w') as zf:
                row_idx = 1
                for order in orders:
                    cur_ws.cell(row=row_idx+2, column=1, value=str(row_idx))
                    cur_ws.cell(row=row_idx+2, column=2, value=order.get('book_name'))
                    cur_ws.cell(row=row_idx+2, column=3, value=order.get('book_goods'))
                    cur_ws.cell(row=row_idx+2, column=4, value=order.get('address'))
                    cur_ws.row_dimensions[row_idx+2].height = 40
                    row_idx = row_idx + 1
                meter_wb.save(xls_filepath)
                zf.write(xls_filepath, xls_filename, compress_type=zipfile.ZIP_DEFLATED)
            os.remove(xls_filepath)
            return zip_filepath, zip_filename
        return None, None
